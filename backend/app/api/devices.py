from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session, joinedload, subqueryload

from app.database import get_db
from app.models.device import Device, DeviceStatus
from app.models.inventory import DeviceOS, DeviceCPU, DeviceRAM, DeviceSoftware, DeviceService
from app.models.tracking import HardwareChange
from app.models.location import Room, Sector, Branch, Company, Unit
from app.models.user import User, UserRole
from app.schemas.device import DeviceResponse, DeviceDetailResponse, DeviceUpdate, HardwareChangeResponse
from app.schemas.device import SoftwareInfo, ServiceInfo
from app.services.audit_service import log_action
from app.utils.security import get_current_user, require_role

router = APIRouter(prefix="/api/devices", tags=["Devices"])


def _build_location_path(db: Session, room_id: int | None) -> str | None:
    if not room_id:
        return None
    sector = db.query(Sector).filter(Sector.id == room_id).first()
    if not sector:
        return None
    branch = db.query(Branch).filter(Branch.id == sector.branch_id).first()
    company = db.query(Company).filter(Company.id == branch.company_id).first() if branch else None
    parts = [p.name for p in [company, branch, sector] if p]
    return " > ".join(parts)


@router.get("/", response_model=list[DeviceResponse])
def list_devices(
    status: DeviceStatus | None = None,
    search: str | None = None,
    unit_id: int | None = None,
    sector_id: int | None = None,
    os_name: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    query = db.query(Device)
    if status:
        query = query.filter(Device.status == status)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            Device.hostname.ilike(pattern) | Device.current_user.ilike(pattern) | Device.agent_id.ilike(pattern)
        )
    if unit_id:
        query = (query.join(Room, Device.room_id == Room.id)
                 .join(Sector).join(Branch).join(Company)
                 .filter(Company.unit_id == unit_id))
    if sector_id:
        query = query.join(Room, Device.room_id == Room.id).filter(Room.sector_id == sector_id)
    if os_name:
        query = query.join(DeviceOS, Device.id == DeviceOS.device_id).filter(DeviceOS.name.ilike(f"%{os_name}%"))

    devices = (query
               .options(
                   subqueryload(Device.os_info),
                   subqueryload(Device.cpus),
                   subqueryload(Device.ram),
               )
               .order_by(Device.hostname)
               .offset((page - 1) * page_size).limit(page_size).all())
    result = []
    for d in devices:
        resp = DeviceResponse.model_validate(d)
        resp.location_path = _build_location_path(db, d.room_id)
        if d.os_info:
            resp.os_name = d.os_info.name
        cpu = next((c for c in d.cpus if c.model), d.cpus[0] if d.cpus else None)
        if cpu:
            resp.cpu_model = cpu.model
        if d.ram:
            resp.ram_total_gb = d.ram.total_gb
        result.append(resp)
    return result


@router.get("/export")
def export_devices(db: Session = Depends(get_db), _=Depends(get_current_user)):
    devices = (db.query(Device)
               .options(
                   subqueryload(Device.os_info),
                   subqueryload(Device.cpus),
                   subqueryload(Device.ram),
                   subqueryload(Device.storage),
                   subqueryload(Device.networks),
                   subqueryload(Device.motherboard),
                   subqueryload(Device.bios),
               )
               .order_by(Device.hostname).all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Dispositivos"

    headers = [
        "Hostname", "Usuário", "Domínio", "Status", "Última Comunicação",
        "Localização", "Sistema Operacional", "Versão OS", "CPU",
        "Núcleos", "Threads", "RAM Total (GB)", "RAM Usada (GB)",
        "Armazenamento", "IP", "MAC", "Placa-mãe", "BIOS",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2B5797", end_color="2B5797", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, device in enumerate(devices, 2):
        location = _build_location_path(db, device.room_id)
        os_info = device.os_info
        cpu = device.cpus[0] if device.cpus else None
        ram = device.ram
        net = next((n for n in device.networks if n.ip_address and not n.ip_address.startswith("127.")), None)
        mb = device.motherboard
        bios = device.bios

        storage_parts = []
        for s in device.storage:
            parts = []
            if s.model:
                parts.append(s.model)
            if s.capacity_gb is not None:
                parts.append(f"{s.capacity_gb:.0f}GB")
            storage_parts.append(" ".join(parts) if parts else "")
        storage_str = "; ".join(storage_parts)

        row_data = [
            device.hostname,
            device.current_user,
            device.domain,
            device.status.value if device.status else "",
            device.last_seen.strftime("%d/%m/%Y %H:%M") if device.last_seen else "",
            location or "",
            os_info.name if os_info else "",
            os_info.version if os_info else "",
            f"{cpu.manufacturer} {cpu.model}".strip() if cpu else "",
            cpu.cores if cpu else "",
            cpu.threads if cpu else "",
            ram.total_gb if ram else "",
            ram.used_gb if ram else "",
            storage_str,
            net.ip_address if net else "",
            net.mac_address if net else "",
            f"{mb.manufacturer} {mb.model}".strip() if mb else "",
            f"{bios.manufacturer} {bios.version}".strip() if bios else "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    for col_idx in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=dispositivos.xlsx"},
    )


@router.get("/{device_id}", response_model=DeviceDetailResponse)
def get_device(device_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    device = (db.query(Device)
              .options(
                  subqueryload(Device.os_info),
                  subqueryload(Device.cpus),
                  subqueryload(Device.ram),
                  subqueryload(Device.ram_slots),
                  subqueryload(Device.storage),
                  subqueryload(Device.networks),
                  subqueryload(Device.motherboard),
                  subqueryload(Device.bios),
                  subqueryload(Device.monitors),
                  subqueryload(Device.printers),
                  subqueryload(Device.local_users),
              )
              .filter(Device.id == device_id).first())
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    resp = DeviceDetailResponse.model_validate(device)
    resp.location_path = _build_location_path(db, device.room_id)
    return resp


@router.put("/{device_id}", response_model=DeviceResponse)
def update_device(
    device_id: int, data: DeviceUpdate, request: Request, db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.TECHNICIAN)),
):
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(device, field, value)
    db.commit()
    db.refresh(device)
    log_action(db, "device_updated", user_id=current_user.id, target_type="device",
               target_id=device_id, details=update_data,
               ip_address=request.client.host if request.client else None)
    return device


@router.delete("/{device_id}", status_code=204)
def delete_device(
    device_id: int, request: Request, db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    device = db.query(Device).filter(Device.id == device_id).first()
    if not device:
        raise HTTPException(status_code=404, detail="Device not found")
    db.delete(device)
    db.commit()
    log_action(db, "device_deleted", user_id=current_user.id, target_type="device",
               target_id=device_id, ip_address=request.client.host if request.client else None)


@router.get("/{device_id}/changes", response_model=list[HardwareChangeResponse])
def get_device_changes(
    device_id: int, page: int = Query(1, ge=1), page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db), _=Depends(get_current_user),
):
    return (db.query(HardwareChange)
            .filter(HardwareChange.device_id == device_id)
            .order_by(HardwareChange.detected_at.desc())
            .offset((page - 1) * page_size).limit(page_size).all())


@router.get("/{device_id}/software", response_model=list[SoftwareInfo])
def get_device_software(device_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    items = db.query(DeviceSoftware).filter(DeviceSoftware.device_id == device_id).order_by(DeviceSoftware.name).all()
    return [SoftwareInfo.model_validate(s) for s in items]


@router.get("/{device_id}/services", response_model=list[ServiceInfo])
def get_device_services(device_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    items = db.query(DeviceService).filter(DeviceService.device_id == device_id).order_by(DeviceService.name).all()
    return [ServiceInfo.model_validate(s) for s in items]
