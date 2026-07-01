import subprocess
from datetime import datetime, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pydantic import BaseModel
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.config import TIMEZONE_BR
from app.database import get_db
from app.models.printer import Printer, PrinterCounter, TonerChange, TonerStock, TonerStockLog, PrinterCollectionSchedule, PrinterTicket
from app.models.user import User, UserRole
from app.services.audit_service import log_action
from app.utils.excel import xlsx_safe
from app.utils.security import get_current_user, require_role

router = APIRouter(prefix="/api/printers", tags=["Printers"])


def _date_trunc(db: Session, column, fmt: str):
    """Cross-database date formatting: fmt uses strftime-style tokens (%Y, %m, %d)."""
    dialect = db.bind.dialect.name
    if dialect == "mysql":
        # MySQL/MariaDB DATE_FORMAT uses the same %Y/%m/%d tokens as strftime
        return func.date_format(column, fmt)
    return func.strftime(fmt, column)


# ── Schemas ──

class PrinterCreate(BaseModel):
    name: str
    model: str | None = None
    serial_number: str | None = None
    ip_address: str | None = None
    location: str | None = None
    sector: str | None = None
    printer_number: str | None = None
    snmp_community: str = "public"
    snmp_version: str = "v2c"
    initial_counter: int = 0


class PrinterUpdate(BaseModel):
    name: str | None = None
    model: str | None = None
    serial_number: str | None = None
    ip_address: str | None = None
    location: str | None = None
    sector: str | None = None
    printer_number: str | None = None
    snmp_community: str | None = None
    snmp_version: str | None = None
    initial_counter: int | None = None
    is_active: bool | None = None


class TonerChangeCreate(BaseModel):
    toner_model: str
    pages_at_change: int | None = None
    notes: str | None = None


class TicketCreate(BaseModel):
    opened_by: str
    description: str


class TicketClose(BaseModel):
    closed_by: str
    resolution: str | None = None


class TicketUpdate(BaseModel):
    status: str | None = None
    resolution: str | None = None
    closed_by: str | None = None


class TonerStockCreate(BaseModel):
    toner_model: str
    quantity: int
    min_quantity: int = 2


class TonerStockAdjust(BaseModel):
    toner_model: str
    quantity: int
    notes: str | None = None


# ── SNMP Helper (raw UDP - no external dependencies) ──

import socket
import struct

SNMP_PAGE_OIDS = [
    "1.3.6.1.2.1.43.10.2.1.4.1.1",
    "1.3.6.1.4.1.1347.42.2.1.1.1.6.1.1",
]

SNMP_TONER_OID = "1.3.6.1.2.1.43.11.1.1.9.1.1"


def _encode_oid(oid_str: str) -> bytes:
    parts = [int(p) for p in oid_str.split(".")]
    encoded = bytes([40 * parts[0] + parts[1]])
    for p in parts[2:]:
        if p < 128:
            encoded += bytes([p])
        else:
            chunks = []
            while p > 0:
                chunks.append(p & 0x7F)
                p >>= 7
            chunks.reverse()
            for i in range(len(chunks) - 1):
                chunks[i] |= 0x80
            encoded += bytes(chunks)
    return encoded


def _encode_length(length: int) -> bytes:
    if length < 128:
        return bytes([length])
    length_bytes = []
    temp = length
    while temp > 0:
        length_bytes.insert(0, temp & 0xFF)
        temp >>= 8
    return bytes([0x80 | len(length_bytes)] + length_bytes)


def _build_snmp_get(community: str, oid_str: str, version: str = "v2c") -> bytes:
    ver = 1 if version == "v2c" else 0
    oid_encoded = _encode_oid(oid_str)

    # NULL value
    null_val = b'\x05\x00'
    # OID TLV
    oid_tlv = b'\x06' + _encode_length(len(oid_encoded)) + oid_encoded
    # VarBind sequence
    varbind = b'\x30' + _encode_length(len(oid_tlv) + len(null_val)) + oid_tlv + null_val
    # VarBindList
    varbind_list = b'\x30' + _encode_length(len(varbind)) + varbind
    # Request ID
    request_id = b'\x02\x04' + struct.pack('>I', 1)
    error_status = b'\x02\x01\x00'
    error_index = b'\x02\x01\x00'
    # GetRequest PDU
    pdu_content = request_id + error_status + error_index + varbind_list
    pdu = b'\xa0' + _encode_length(len(pdu_content)) + pdu_content
    # Version
    version_tlv = b'\x02\x01' + bytes([ver])
    # Community
    comm_bytes = community.encode()
    community_tlv = b'\x04' + _encode_length(len(comm_bytes)) + comm_bytes
    # Message
    msg_content = version_tlv + community_tlv + pdu
    message = b'\x30' + _encode_length(len(msg_content)) + msg_content
    return message


def _parse_snmp_response(data: bytes) -> str | None:
    try:
        i = 0
        while i < len(data):
            if data[i] == 0x02 and i + 20 < len(data):
                # Look for integer values (counter)
                pass
            i += 1

        # Simple approach: find the last value in the response
        # Walk backwards through TLV structures
        pos = len(data) - 1
        while pos > 10:
            # Check for Integer (0x02), Counter32 (0x41), Gauge32 (0x42)
            if data[pos - 1] == 0x02 or data[pos - 1] == 0x41 or data[pos - 1] == 0x42:
                break
            # Check for OctetString (0x04)
            if data[pos - 1] == 0x04:
                break
            pos -= 1

        # Find the last varbind value
        idx = data.rfind(b'\x05\x00')  # NULL marker
        if idx == -1:
            # No NULL, try to find value after OID
            # Parse from the response PDU
            pass

        # More robust: parse the ASN.1 structure
        def parse_tlv(d: bytes, offset: int) -> tuple:
            tag = d[offset]
            offset += 1
            length = d[offset]
            offset += 1
            if length & 0x80:
                num_bytes = length & 0x7F
                length = int.from_bytes(d[offset:offset + num_bytes], 'big')
                offset += num_bytes
            return tag, length, offset

        # Skip outer SEQUENCE
        _, _, pos = parse_tlv(data, 0)
        # Skip version
        _, vlen, pos = parse_tlv(data, pos)
        pos += vlen
        # Skip community
        _, clen, pos = parse_tlv(data, pos)
        pos += clen
        # GetResponse PDU
        _, _, pos = parse_tlv(data, pos)
        # Request ID
        _, rlen, pos = parse_tlv(data, pos)
        pos += rlen
        # Error status
        _, elen, pos = parse_tlv(data, pos)
        error_val = int.from_bytes(data[pos:pos + elen], 'big')
        pos += elen
        if error_val != 0:
            return None
        # Error index
        _, eilen, pos = parse_tlv(data, pos)
        pos += eilen
        # VarBindList SEQUENCE
        _, _, pos = parse_tlv(data, pos)
        # VarBind SEQUENCE
        _, _, pos = parse_tlv(data, pos)
        # OID
        _, olen, pos = parse_tlv(data, pos)
        pos += olen
        # Value
        val_tag, val_len, pos = parse_tlv(data, pos)
        val_data = data[pos:pos + val_len]

        if val_tag in (0x02, 0x41, 0x42, 0x46):  # Integer, Counter32, Gauge32, Counter64
            return str(int.from_bytes(val_data, 'big'))
        elif val_tag == 0x04:  # OctetString
            try:
                return val_data.decode('utf-8').strip()
            except Exception:
                return val_data.hex()
        elif val_tag in (0x80, 0x81, 0x82):  # noSuchObject/Instance/endOfMibView
            return None

        return str(int.from_bytes(val_data, 'big')) if val_data else None

    except Exception:
        return None


def _snmp_get(ip: str, oid: str, community: str = "public", version: str = "v2c") -> str | None:
    try:
        packet = _build_snmp_get(community, oid, version)
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.settimeout(3)
        sock.sendto(packet, (ip, 161))
        data, _ = sock.recvfrom(4096)
        sock.close()
        return _parse_snmp_response(data)
    except socket.timeout:
        return None
    except Exception:
        return None


def _collect_snmp_counter(printer: Printer) -> int | None:
    if not printer.ip_address:
        return None
    for oid in SNMP_PAGE_OIDS:
        val = _snmp_get(printer.ip_address, oid, printer.snmp_community, printer.snmp_version)
        if val and val.isdigit():
            return int(val)
    return None


def _collect_snmp_toner(printer: Printer) -> int | None:
    if not printer.ip_address:
        return None
    val = _snmp_get(printer.ip_address, SNMP_TONER_OID, printer.snmp_community, printer.snmp_version)
    if val and val.lstrip("-").isdigit():
        return int(val)
    return None


# ── Dashboard Stats ──

@router.get("/dashboard-stats")
def get_printer_dashboard(db: Session = Depends(get_db), _=Depends(get_current_user)):
    printers = db.query(Printer).filter(Printer.is_active == True).all()

    total_pages = 0
    total_toner_changes = 0
    pages_per_printer = []
    toner_per_printer = []

    for p in printers:
        last_counter = (db.query(PrinterCounter).filter(PrinterCounter.printer_id == p.id)
                        .order_by(PrinterCounter.collected_at.desc()).first())
        effective = max((last_counter.total_pages if last_counter else 0) - p.initial_counter, 0)
        total_pages += effective

        toner_count = db.query(func.count(TonerChange.id)).filter(TonerChange.printer_id == p.id).scalar() or 0
        total_toner_changes += toner_count

        pages_per_printer.append({"name": p.name, "pages": effective, "id": p.id})
        toner_per_printer.append({"name": p.name, "toner_changes": toner_count, "id": p.id})

    # Toner models consumed
    toner_models = (db.query(TonerChange.toner_model, func.count(TonerChange.id))
                    .group_by(TonerChange.toner_model)
                    .order_by(func.count(TonerChange.id).desc()).all())
    toner_by_model = [{"model": m, "count": c} for m, c in toner_models]

    # Monthly pages (last 12 months from counters)
    month_expr = _date_trunc(db, PrinterCounter.collected_at, '%Y-%m')
    monthly = (db.query(
        month_expr.label('month'),
        func.max(PrinterCounter.total_pages).label('max_pages'),
        PrinterCounter.printer_id,
    ).group_by(month_expr, PrinterCounter.printer_id)
     .order_by(month_expr).all())

    months_data: dict[str, int] = {}
    prev_by_printer: dict[int, int] = {}
    for month, max_pages, pid in monthly:
        if pid in prev_by_printer:
            diff = max(max_pages - prev_by_printer[pid], 0)
            months_data[month] = months_data.get(month, 0) + diff
        prev_by_printer[pid] = max_pages

    monthly_chart = [{"month": m, "pages": p} for m, p in sorted(months_data.items())]

    # Stock summary
    stocks = db.query(TonerStock).all()
    total_stock = sum(s.quantity for s in stocks)
    low_stock_count = sum(1 for s in stocks if s.quantity <= s.min_quantity)

    # ── Consumption by period (daily, weekly, monthly) ──
    now = datetime.now(TIMEZONE_BR)

    def _pages_in_period(since: datetime) -> list[dict]:
        day_expr = _date_trunc(db, PrinterCounter.collected_at, '%Y-%m-%d')
        rows = (db.query(
            day_expr.label('day'),
            func.max(PrinterCounter.total_pages),
            func.min(PrinterCounter.total_pages),
            PrinterCounter.printer_id,
        ).filter(PrinterCounter.collected_at >= since)
         .group_by(day_expr, PrinterCounter.printer_id)
         .order_by(day_expr).all())

        daily: dict[str, int] = {}
        for day, max_p, min_p, pid in rows:
            daily[day] = daily.get(day, 0) + max(max_p - min_p, 0)
        return [{"date": d, "pages": p} for d, p in sorted(daily.items())]

    def _toners_in_period(since: datetime) -> list[dict]:
        day_expr = _date_trunc(db, TonerChange.changed_at, '%Y-%m-%d')
        rows = (db.query(
            day_expr.label('day'),
            func.count(TonerChange.id),
        ).filter(TonerChange.changed_at >= since)
         .group_by(day_expr).order_by(day_expr).all())
        return [{"date": d, "count": c} for d, c in rows]

    daily_pages = _pages_in_period(now - timedelta(days=7))
    weekly_pages = _pages_in_period(now - timedelta(days=30))
    monthly_pages_detail = _pages_in_period(now - timedelta(days=90))

    daily_toners = _toners_in_period(now - timedelta(days=7))
    weekly_toners = _toners_in_period(now - timedelta(days=30))
    monthly_toners = _toners_in_period(now - timedelta(days=90))

    # Summary cards
    pages_today = sum(r["pages"] for r in daily_pages if r["date"] == now.strftime('%Y-%m-%d'))
    pages_week = sum(r["pages"] for r in daily_pages)
    pages_month = sum(r["pages"] for r in weekly_pages)

    toners_today = sum(r["count"] for r in daily_toners if r["date"] == now.strftime('%Y-%m-%d'))
    toners_week = sum(r["count"] for r in daily_toners)
    toners_month = sum(r["count"] for r in weekly_toners)

    open_tickets = db.query(func.count(PrinterTicket.id)).filter(PrinterTicket.status == "aberto").scalar() or 0
    total_tickets = db.query(func.count(PrinterTicket.id)).scalar() or 0

    return {
        "total_printers": len(printers),
        "total_pages": total_pages,
        "total_toner_changes": total_toner_changes,
        "total_stock": total_stock,
        "low_stock_count": low_stock_count,
        "open_tickets": open_tickets,
        "total_tickets": total_tickets,
        "pages_per_printer": sorted(pages_per_printer, key=lambda x: x["pages"], reverse=True),
        "toner_per_printer": sorted(toner_per_printer, key=lambda x: x["toner_changes"], reverse=True),
        "toner_by_model": toner_by_model,
        "monthly_pages": monthly_chart,
        "pages_today": pages_today,
        "pages_week": pages_week,
        "pages_month": pages_month,
        "toners_today": toners_today,
        "toners_week": toners_week,
        "toners_month": toners_month,
        "daily_pages_chart": daily_pages,
        "weekly_pages_chart": weekly_pages,
        "monthly_pages_chart": monthly_pages_detail,
        "daily_toners_chart": daily_toners,
        "weekly_toners_chart": weekly_toners,
        "monthly_toners_chart": monthly_toners,
    }


# ── Scheduled (automatic) collection ──

class ScheduleUpdate(BaseModel):
    enabled: bool
    interval_minutes: int


@router.get("/schedule")
def get_schedule(db: Session = Depends(get_db), _=Depends(get_current_user)):
    sched = db.query(PrinterCollectionSchedule).first()
    if not sched:
        sched = PrinterCollectionSchedule(enabled=False, interval_minutes=60)
        db.add(sched)
        db.commit()
        db.refresh(sched)
    return {
        "enabled": sched.enabled,
        "interval_minutes": sched.interval_minutes,
        "last_run": sched.last_run.isoformat() if sched.last_run else None,
    }


@router.put("/schedule")
def update_schedule(
    data: ScheduleUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.TECHNICIAN)),
):
    if data.interval_minutes < 5:
        raise HTTPException(400, "Intervalo mínimo é de 5 minutos")
    sched = db.query(PrinterCollectionSchedule).first()
    if not sched:
        sched = PrinterCollectionSchedule()
        db.add(sched)
    sched.enabled = data.enabled
    sched.interval_minutes = data.interval_minutes
    db.commit()

    log_action(db, "printer_schedule_updated", user_id=current_user.id,
               details={"enabled": data.enabled, "interval_minutes": data.interval_minutes})

    return {"message": "Agendamento atualizado", "enabled": sched.enabled, "interval_minutes": sched.interval_minutes}


# ── Ping ──

@router.post("/{printer_id}/ping")
def ping_printer(
    printer_id: int,
    db: Session = Depends(get_db), _=Depends(get_current_user),
):
    printer = db.query(Printer).filter(Printer.id == printer_id).first()
    if not printer:
        raise HTTPException(404, "Impressora não encontrada")
    if not printer.ip_address:
        raise HTTPException(400, "Impressora sem IP configurado")
    try:
        result = subprocess.run(
            ["ping", "-c", "2", "-W", "2", printer.ip_address],
            capture_output=True, text=True, timeout=10,
        )
        if result.returncode != 0:
            result = subprocess.run(
                ["ping", "-n", "2", "-w", "2000", printer.ip_address],
                capture_output=True, text=True, timeout=10,
            )
        online = result.returncode == 0
        return {"ip": printer.ip_address, "online": online, "output": result.stdout[-300:] if result.stdout else result.stderr[-300:]}
    except Exception as e:
        return {"ip": printer.ip_address, "online": False, "output": str(e)}


# ── Printer CRUD ──

@router.get("/")
def list_printers(
    active_only: bool = True,
    db: Session = Depends(get_db), _=Depends(get_current_user),
):
    query = db.query(Printer)
    if active_only:
        query = query.filter(Printer.is_active == True)
    printers = query.order_by(Printer.name).all()

    result = []
    for p in printers:
        last_counter = (db.query(PrinterCounter)
                        .filter(PrinterCounter.printer_id == p.id)
                        .order_by(PrinterCounter.collected_at.desc()).first())
        toner_count = db.query(func.count(TonerChange.id)).filter(TonerChange.printer_id == p.id).scalar() or 0
        effective_pages = (last_counter.total_pages - p.initial_counter) if last_counter else 0
        result.append({
            "id": p.id, "name": p.name, "model": p.model, "serial_number": p.serial_number,
            "ip_address": p.ip_address, "location": p.location, "sector": p.sector,
            "printer_number": p.printer_number, "snmp_version": p.snmp_version,
            "initial_counter": p.initial_counter, "is_active": p.is_active,
            "last_page_count": last_counter.total_pages if last_counter else None,
            "effective_pages": max(effective_pages, 0),
            "last_collected": last_counter.collected_at.isoformat() if last_counter else None,
            "toner_changes": toner_count,
        })
    return result


@router.post("/")
def create_printer(
    data: PrinterCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.TECHNICIAN)),
):
    printer = Printer(**data.model_dump())
    db.add(printer)
    db.commit()
    db.refresh(printer)
    return {"id": printer.id, "name": printer.name, "message": "Impressora cadastrada"}


@router.put("/{printer_id}")
def update_printer(
    printer_id: int, data: PrinterUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.TECHNICIAN)),
):
    printer = db.query(Printer).filter(Printer.id == printer_id).first()
    if not printer:
        raise HTTPException(404, "Impressora não encontrada")
    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(printer, field, value)
    db.commit()
    return {"message": "Impressora atualizada"}


@router.delete("/{printer_id}", status_code=204)
def delete_printer(
    printer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    printer = db.query(Printer).filter(Printer.id == printer_id).first()
    if not printer:
        raise HTTPException(404, "Impressora não encontrada")
    db.delete(printer)
    db.commit()


# ── SNMP Collection ──

@router.post("/{printer_id}/collect")
def collect_counter(
    printer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.TECHNICIAN)),
):
    printer = db.query(Printer).filter(Printer.id == printer_id).first()
    if not printer:
        raise HTTPException(404, "Impressora não encontrada")

    pages = _collect_snmp_counter(printer)
    if pages is None:
        raise HTTPException(400, "Falha ao coletar via SNMP. Verifique IP e comunidade.")

    counter = PrinterCounter(printer_id=printer.id, total_pages=pages, source="snmp")
    db.add(counter)
    db.commit()

    effective = max(pages - printer.initial_counter, 0)
    return {"total_pages": pages, "effective_pages": effective, "initial_counter": printer.initial_counter}


def collect_all_printers(db: Session) -> dict:
    printers = db.query(Printer).filter(Printer.is_active == True, Printer.ip_address != None).all()
    results = []
    for p in printers:
        pages = _collect_snmp_counter(p)
        if pages is not None:
            counter = PrinterCounter(printer_id=p.id, total_pages=pages, source="snmp")
            db.add(counter)
            results.append({"printer": p.name, "pages": pages, "status": "ok"})
        else:
            results.append({"printer": p.name, "pages": None, "status": "failed"})
    db.commit()
    return {"results": results, "collected": sum(1 for r in results if r["status"] == "ok")}


@router.post("/collect-all")
def collect_all_counters(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.TECHNICIAN)),
):
    return collect_all_printers(db)


# ── Scheduled (automatic) collection ──

@router.get("/{printer_id}/history")
def get_counter_history(
    printer_id: int,
    limit: int = Query(100, ge=1, le=1000),
    db: Session = Depends(get_db), _=Depends(get_current_user),
):
    printer = db.query(Printer).filter(Printer.id == printer_id).first()
    if not printer:
        raise HTTPException(404, "Impressora não encontrada")
    counters = (db.query(PrinterCounter)
                .filter(PrinterCounter.printer_id == printer_id)
                .order_by(PrinterCounter.collected_at.desc()).limit(limit).all())
    return [{
        "id": c.id, "total_pages": c.total_pages,
        "effective_pages": max(c.total_pages - printer.initial_counter, 0),
        "collected_at": c.collected_at.isoformat(), "source": c.source,
    } for c in counters]


# ── Rankings ──

@router.get("/ranking")
def get_printer_ranking(
    top: int = Query(10, ge=5, le=50),
    db: Session = Depends(get_db), _=Depends(get_current_user),
):
    printers = db.query(Printer).filter(Printer.is_active == True).all()
    data = []
    for p in printers:
        last = (db.query(PrinterCounter).filter(PrinterCounter.printer_id == p.id)
                .order_by(PrinterCounter.collected_at.desc()).first())
        effective = max((last.total_pages if last else 0) - p.initial_counter, 0)
        data.append({
            "id": p.id, "name": p.name, "model": p.model,
            "location": p.location, "sector": p.sector,
            "effective_pages": effective,
        })

    by_pages = sorted(data, key=lambda x: x["effective_pages"], reverse=True)
    return {
        "most": by_pages[:top],
        "least": list(reversed(by_pages[-top:])) if len(by_pages) >= top else list(reversed(by_pages)),
    }


# ── Toner Changes ──

@router.post("/{printer_id}/toner-change")
def register_toner_change(
    printer_id: int, data: TonerChangeCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.TECHNICIAN)),
):
    printer = db.query(Printer).filter(Printer.id == printer_id).first()
    if not printer:
        raise HTTPException(404, "Impressora não encontrada")

    change = TonerChange(
        printer_id=printer.id, toner_model=data.toner_model,
        changed_by=current_user.id, pages_at_change=data.pages_at_change,
        notes=data.notes,
    )
    db.add(change)

    stock = db.query(TonerStock).filter(TonerStock.toner_model == data.toner_model).first()
    if stock and stock.quantity > 0:
        stock.quantity -= 1
        log = TonerStockLog(
            toner_model=data.toner_model, action="substituição", quantity=1,
            user_id=current_user.id, printer_id=printer.id,
            notes=f"Substituição em {printer.name}",
        )
        db.add(log)

    db.commit()

    log_action(db, "toner_changed", user_id=current_user.id, target_type="printer",
               target_id=printer_id, details={"toner_model": data.toner_model})

    return {"message": "Toner substituído", "stock_remaining": stock.quantity if stock else None}


@router.get("/{printer_id}/toner-history")
def get_toner_history(
    printer_id: int,
    db: Session = Depends(get_db), _=Depends(get_current_user),
):
    changes = (db.query(TonerChange)
               .filter(TonerChange.printer_id == printer_id)
               .order_by(TonerChange.changed_at.desc()).all())
    return [{
        "id": c.id, "toner_model": c.toner_model,
        "changed_by": c.user.username if c.user else "-",
        "pages_at_change": c.pages_at_change,
        "notes": c.notes,
        "changed_at": c.changed_at.isoformat(),
    } for c in changes]


# ── Toner Stock ──

@router.get("/stock")
def list_stock(db: Session = Depends(get_db), _=Depends(get_current_user)):
    stocks = db.query(TonerStock).order_by(TonerStock.toner_model).all()
    return [{
        "id": s.id, "toner_model": s.toner_model,
        "quantity": s.quantity, "min_quantity": s.min_quantity,
        "low_stock": s.quantity <= s.min_quantity,
    } for s in stocks]


@router.post("/stock")
def create_or_update_stock(
    data: TonerStockCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.TECHNICIAN)),
):
    stock = db.query(TonerStock).filter(TonerStock.toner_model == data.toner_model).first()
    if stock:
        stock.quantity = data.quantity
        stock.min_quantity = data.min_quantity
    else:
        stock = TonerStock(**data.model_dump())
        db.add(stock)

    log = TonerStockLog(
        toner_model=data.toner_model, action="cadastro", quantity=data.quantity,
        user_id=current_user.id, notes="Cadastro/atualização de estoque",
    )
    db.add(log)
    db.commit()
    return {"message": "Estoque atualizado", "quantity": stock.quantity}


@router.post("/stock/restock")
def restock_toner(
    data: TonerStockAdjust,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.TECHNICIAN)),
):
    stock = db.query(TonerStock).filter(TonerStock.toner_model == data.toner_model).first()
    if not stock:
        raise HTTPException(404, "Modelo de toner não encontrado no estoque")
    stock.quantity += data.quantity

    log = TonerStockLog(
        toner_model=data.toner_model, action="reabastecimento", quantity=data.quantity,
        user_id=current_user.id, notes=data.notes or "Reabastecimento",
    )
    db.add(log)
    db.commit()
    return {"message": f"Adicionadas {data.quantity} unidades", "new_quantity": stock.quantity}


@router.get("/stock/logs")
def get_stock_logs(
    toner_model: str | None = None,
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db), _=Depends(get_current_user),
):
    query = db.query(TonerStockLog)
    if toner_model:
        query = query.filter(TonerStockLog.toner_model == toner_model)
    logs = query.order_by(TonerStockLog.created_at.desc()).limit(limit).all()
    return [{
        "id": l.id, "toner_model": l.toner_model, "action": l.action,
        "quantity": l.quantity, "user": l.user.username if l.user else "-",
        "printer": l.printer.name if l.printer else None,
        "printer_number": l.printer.printer_number if l.printer else None,
        "printer_location": l.printer.location if l.printer else None,
        "printer_sector": l.printer.sector if l.printer else None,
        "notes": l.notes, "created_at": l.created_at.isoformat(),
    } for l in logs]


@router.get("/stock/export")
def export_toner_logs(
    db: Session = Depends(get_db), _=Depends(get_current_user),
):
    logs = (db.query(TonerStockLog)
            .order_by(TonerStockLog.created_at.desc()).all())

    wb = Workbook()
    ws = wb.active
    ws.title = "Dispensação de Toner"

    headers = ["Data/Hora", "Ação", "Modelo Toner", "Qtd", "Impressora", "Nº Impressora",
               "Local", "Setor", "Operador", "Observações"]
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
    thin = Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"), bottom=Side(style="thin"))

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for ri, l in enumerate(logs, 2):
        row = [
            l.created_at.strftime("%d/%m/%Y %H:%M") if l.created_at else "",
            l.action or "",
            l.toner_model or "",
            l.quantity,
            l.printer.name if l.printer else "",
            l.printer.printer_number if l.printer else "",
            l.printer.location if l.printer else "",
            l.printer.sector if l.printer else "",
            l.user.username if l.user else "",
            l.notes or "",
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=xlsx_safe(val))
            cell.border = thin

    for ci in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=ci).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 3, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=dispensacao_toner_{datetime.now().strftime('%Y%m%d')}.xlsx"},
    )


# ── Chamados Técnicos ──

def _ticket_row(t: PrinterTicket) -> dict:
    p = t.printer
    return {
        "id": t.id,
        "os_number": t.os_number,
        "printer_id": t.printer_id,
        "printer_name": p.name if p else "",
        "printer_location": p.location if p else "",
        "printer_sector": p.sector if p else "",
        "printer_number": p.printer_number if p else "",
        "opened_by": t.opened_by,
        "description": t.description,
        "status": t.status,
        "created_at": t.created_at.isoformat() if t.created_at else None,
        "closed_at": t.closed_at.isoformat() if t.closed_at else None,
        "closed_by": t.closed_by,
        "resolution": t.resolution,
    }


@router.get("/tickets")
def list_tickets(
    printer_id: int | None = Query(default=None),
    status: str | None = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = db.query(PrinterTicket).join(Printer)
    if printer_id:
        q = q.filter(PrinterTicket.printer_id == printer_id)
    if status:
        q = q.filter(PrinterTicket.status == status)
    q = q.order_by(PrinterTicket.created_at.desc())
    return [_ticket_row(t) for t in q.all()]


@router.post("/{printer_id}/tickets")
def create_ticket(
    printer_id: int,
    data: TicketCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    printer = db.query(Printer).filter(Printer.id == printer_id).first()
    if not printer:
        raise HTTPException(404, "Impressora não encontrada")

    ticket = PrinterTicket(
        os_number="TEMP",
        printer_id=printer_id,
        opened_by=data.opened_by,
        description=data.description,
        status="aberto",
    )
    db.add(ticket)
    db.flush()

    year = datetime.now(TIMEZONE_BR).year
    ticket.os_number = f"OS-{year}-{ticket.id:04d}"
    db.commit()
    db.refresh(ticket)

    log_action(db, current_user.id, "create_ticket", "printer_ticket", ticket.id,
               {"os_number": ticket.os_number, "printer_id": printer_id})

    return _ticket_row(ticket)


@router.put("/tickets/{ticket_id}")
def update_ticket(
    ticket_id: int,
    data: TicketUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ticket = db.query(PrinterTicket).filter(PrinterTicket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(404, "Chamado não encontrado")

    if data.status is not None:
        ticket.status = data.status
        if data.status == "fechado" and not ticket.closed_at:
            ticket.closed_at = datetime.now(TIMEZONE_BR)
        elif data.status == "aberto":
            ticket.closed_at = None
            ticket.closed_by = None
    if data.resolution is not None:
        ticket.resolution = data.resolution
    if data.closed_by is not None:
        ticket.closed_by = data.closed_by

    db.commit()
    db.refresh(ticket)
    return _ticket_row(ticket)


@router.delete("/tickets/{ticket_id}")
def delete_ticket(
    ticket_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(UserRole.TECHNICIAN)),
):
    ticket = db.query(PrinterTicket).filter(PrinterTicket.id == ticket_id).first()
    if not ticket:
        raise HTTPException(404, "Chamado não encontrado")
    db.delete(ticket)
    db.commit()
    return {"ok": True}


@router.get("/tickets/export")
def export_tickets(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tickets = (
        db.query(PrinterTicket)
        .join(Printer)
        .order_by(PrinterTicket.created_at.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Chamados Técnicos"

    headers = ["Nº OS", "Impressora", "Local", "Setor", "N.Impressora",
               "Aberto por", "Descrição", "Status",
               "Data Abertura", "Data Fechamento", "Fechado por", "Resolução"]

    hfill = PatternFill("solid", fgColor="1565FF")
    hfont = Font(bold=True, color="FFFFFF")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin

    for ri, t in enumerate(tickets, 2):
        p = t.printer
        row = [
            t.os_number,
            p.name if p else "",
            p.location if p else "",
            p.sector if p else "",
            p.printer_number if p else "",
            t.opened_by,
            t.description,
            t.status,
            t.created_at.strftime("%d/%m/%Y %H:%M") if t.created_at else "",
            t.closed_at.strftime("%d/%m/%Y %H:%M") if t.closed_at else "",
            t.closed_by or "",
            t.resolution or "",
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=xlsx_safe(val))
            cell.border = thin

    for ci in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=ci).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 3, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"chamados_tecnicos_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
